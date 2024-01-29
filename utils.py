import openpyxl
import random
import cmath

def createExcel(filename, sheet1, sheet2):
    """Creates an excel file with name filename and sheet1 for 100 rows of random numbers and sheet2 for the roots"""
    wb = openpyxl.Workbook()
    sheet_1 = wb.active
    sheet_1.title = sheet1
    sheet_2 = wb.create_sheet(title=sheet2)

    sheet_1.append(['A', 'B', 'C'])
    
    for _ in range(100):
        row_data = [random.randint(1, 10), random.randint(0, 10), random.randint(0, 10)]
        sheet_1.append(row_data)

    wb.save(filename)

def findRoots(filename, sheet1, sheet2):
    wb = openpyxl.load_workbook(filename)
    sheet_1 = wb[sheet1]

    roots_list = []

    for row in sheet_1.iter_rows(min_row=2, values_only=True):
        a, b, c = row
        roots = quadratic_roots(a, b, c)

        # Check if roots have imaginary part
        formatted_roots = format_roots(roots)
        roots_list.append(formatted_roots)

    wb.close()
    return roots_list

def loadExcel(filename, sheet1, sheet2, roots_list):
    """Loads excel sheet2 with the roots"""
    wb = openpyxl.load_workbook(filename)
    sheet_2 = wb[sheet2]

    sheet_2.append(['Root 1', 'Root 2'])

    for roots in roots_list:
        root1, root2 = roots

        root1_str = str(root1)
        root2_str = str(root2)

        sheet_2.append([root1_str, root2_str])

    wb.save(filename)

def format_roots(roots):
    formatted_roots = []
    for root in roots:
        if root.imag == 0:
            formatted_roots.append(root.real)
        else:
            formatted_roots.append(root)
    return formatted_roots

def quadratic_roots(a, b, c):
    """Uses the quadratic formula to find roots"""
    discriminant = cmath.sqrt(b**2 - 4*a*c)
    root1 = (-b + discriminant) / (2*a)
    root2 = (-b - discriminant) / (2*a)
    return root1, root2
